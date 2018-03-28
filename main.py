import openpyxl
import time
import json
from datetime import datetime
from constants import *

import tableextraction
import categoriesreader
import exceladapter
import translators

import excelbuilder

# abrir

categories_reader = categoriesreader.CategoriesReader()
actual_excelreader = exceladapter.excelreader.ExcelReader('inputs/PLANILLA CONTABILIDAD ACTIVA 2017.xlsx')
current_accounts_excelreader = exceladapter.ExcelReader('inputs/CUENTAS CORRIENTES.xlsx')

cheques_sheet = actual_excelreader.get_sheet(CHEQUES_SHEET_NAME)
caja_sheet = actual_excelreader.get_sheet(CAJA_SHEET_NAME)
credicoop_sheet = actual_excelreader.get_sheet(CREDICOOP_SHEET_NAME)

cuentas_corrientes_sheet = current_accounts_excelreader.get_sheet(CUENTAS_CORRIENTES_SHEET_NAME)

# initialize lists

actual_flows = []
projected_flows = []
new_categories = set()

for rowNum in range(3, cheques_sheet.max_row):  # skip the first row
    providers_amount = cheques_sheet.cell(row=rowNum, column=6).value

    if not providers_amount:
        continue

    check = {}
    date_value = cheques_sheet.cell(row=rowNum, column=1).value
    details = cheques_sheet.cell(row=rowNum, column=2).value

    check['date'] = date_value.strftime("%d/%m/%Y") if date_value else "N/D"
    check['year'] = date_value.strftime("%Y") if date_value else "N/D"
    check['week'] = int(date_value.strftime("%U")) + 1 if date_value else "N/D"

    check['flow'] = ACTUAL_FLOW
    check['flexibility'] = ""
    check['type'] = CHEQUES_TYPE  # cheques
    check['details'] = details

    category, account = categories_reader.get_category_from_details(details)
    check['category'] = category
    check['account'] = account
    check['income'] = ""
    check['expense'] = translators.to_google_num(providers_amount)

    if date_value and date_value > datetime.now():
        check['flow'] = ESTIMATED_FLOW
        check['flexibility'] = INFLEXIBLE
        projected_flows.append(check)
    else:
        actual_flows.append(check)

table_unpacker = tableextraction.TableUnpacker(caja_sheet)

for rowNum in range(3, caja_sheet.max_row):
    cash_flow = {}

    row_unpacker = table_unpacker.get_row_unpacker(rowNum)

    account = row_unpacker.get_value_at(3)
    details = row_unpacker.get_value_at(2)

    if details == CARGAS_SOCIALES:
        account = SUELDOS

    if details in BANKS:
        account = PRESTAMOS_BANCARIOS

    if (account == T_E_CUENTAS_PROPIAS and details == FELSIM_CREDICOOP):
        continue

    expense = row_unpacker.get_value_at(5)
    income = row_unpacker.get_value_at(6)
    category = categories_reader.get_category_from_account_and_details(account, details)

    if category == NEW_CATEGORY:
        new_categories.add('{"account": "%s", "details": "%s"}' % (account, details))

    date_value = row_unpacker.get_value_at(1)

    cash_flow['date'], cash_flow['week'], cash_flow['year'] = translators.unpack_dates(date_value)

    cash_flow['flow'] = ACTUAL_FLOW
    cash_flow['flexibility'] = ""

    cash_flow['type'] = CAJAS_TYPE  # caja

    cash_flow['details'] = details

    cash_flow['category'] = category
    cash_flow['account'] = account
    cash_flow['income'] = ""
    cash_flow['expense'] = translators.to_google_num(expense) if expense else ""
    cash_flow['income'] = translators.to_google_num(income) if income else ""

    # print(cash_flow)

    actual_flows.append(cash_flow)

credicoop_unpacker = tableextraction.TableUnpacker(credicoop_sheet)

for rowNum in range(3, credicoop_sheet.max_row):
    cash_flow = {}

    row_unpacker = credicoop_unpacker.get_row_unpacker(rowNum)

    details = row_unpacker.get_value_at(6)
    account = row_unpacker.get_value_at(7)

    if details == CARGAS_SOCIALES:
        account = SUELDOS

    if details in BANKS:
        account = PRESTAMOS_BANCARIOS

    if details == ANNULED:
        continue

    if account == T_E_CUENTAS_PROPIAS and details == "FELSIM CAJA":
        continue

    if details.startswith(MORATORIA_AFIP):
        account = MORATORIAS

    expense = row_unpacker.get_value_at(9)
    income = row_unpacker.get_value_at(10)
    category = categories_reader.get_category_from_account_and_details(account, details)

    if category == NEW_CATEGORY:
        new_categories.add('{"account": "%s", "details": "%s"}' % (account, details))

    date_value = row_unpacker.get_value_at(1)
    check_clearing_date_value = row_unpacker.get_value_at(4)

    if check_clearing_date_value:
        date_value = check_clearing_date_value

    cash_flow['date'], cash_flow['week'], cash_flow['year'] = translators.unpack_dates(date_value)

    cash_flow['flow'] = ACTUAL_FLOW
    cash_flow['flexibility'] = ""
    cash_flow['type'] = CREDICOOP_TYPE
    cash_flow['details'] = details

    cash_flow['category'] = category
    cash_flow['account'] = account
    cash_flow['income'] = ""
    cash_flow['expense'] = translators.to_google_num(expense) if expense else ""
    cash_flow['income'] = translators.to_google_num(income) if income else ""

    # print(cash_flow)

    actual_flows.append(cash_flow)

# crear excel nuevo
filename = 'outputs/consolidado_real_' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'
actual_excelwriter = exceladapter.ExcelWriter(filename)
new_sheet = actual_excelwriter.create_sheet('Consolidado')

actual_flow_builder = excelbuilder.BasicBuilder(new_sheet, actual_flows)

actual_flow_builder.add_header("A", "Semana")
actual_flow_builder.add_header("B", "Año")
actual_flow_builder.add_header("C", "Flujo")
actual_flow_builder.add_header("D", "Tipo")
actual_flow_builder.add_header("E", "Rubro")
actual_flow_builder.add_header("F", "Fecha")
actual_flow_builder.add_header("G", "Cuenta")
actual_flow_builder.add_header("H", "Detalle")
actual_flow_builder.add_header("I", "Ingreso")
actual_flow_builder.add_header("J", "Egreso")

actual_flow_builder.map_column("A", "week")
actual_flow_builder.map_column("B", "year")
actual_flow_builder.map_column("C", "flow")
actual_flow_builder.map_column("D", "type")
actual_flow_builder.map_column("E", "category")
actual_flow_builder.map_column("F", "date")
actual_flow_builder.map_column("G", "account")
actual_flow_builder.map_column("H", "details")
actual_flow_builder.map_column("I", "income")
actual_flow_builder.map_column("J", "expense")

actual_flow_builder.build()
actual_excelwriter.save()

# crear excel de proyecciones
projected_flow_filename = 'outputs/proyecciones_' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'
projected_excelwriter = exceladapter.ExcelWriter(projected_flow_filename)
projected_sheet = projected_excelwriter.create_sheet('Proyectado')

projected_flow_builder = excelbuilder.BasicBuilder(projected_sheet, projected_flows)

projected_flow_builder.add_header("A", "Semana")
projected_flow_builder.add_header("B", "Año")
projected_flow_builder.add_header("C", "Flujo")
projected_flow_builder.add_header("D", "Tipo")
projected_flow_builder.add_header("E", "Rubro")
projected_flow_builder.add_header("F", "Fecha")
projected_flow_builder.add_header("G", "Cuenta")
projected_flow_builder.add_header("H", "Detalle")
projected_flow_builder.add_header("I", "Ingreso")
projected_flow_builder.add_header("J", "Egreso")

projected_flow_builder.map_column("A", "week")
projected_flow_builder.map_column("B", "year")
projected_flow_builder.map_column("C", "flow")
projected_flow_builder.map_column("D", "type")
projected_flow_builder.map_column("E", "category")
projected_flow_builder.map_column("F", "date")
projected_flow_builder.map_column("G", "account")
projected_flow_builder.map_column("H", "details")
projected_flow_builder.map_column("I", "income")
projected_flow_builder.map_column("J", "expense")

projected_flow_builder.build()
projected_excelwriter.save()

# crear excel de categorías faltantes

missing_categories_filename = 'outputs/rubros_faltantes_' + time.strftime("%Y%m%d-%H%M%S") + '.xlsx'
missing_categories_excelwriter = exceladapter.ExcelWriter(missing_categories_filename)
new_categories_sheet = missing_categories_excelwriter.create_sheet('Rubros Faltantes')

missing_categories_builder = excelbuilder.BasicBuilder(new_categories_sheet, new_categories)

missing_categories_builder.add_header("A", "CuentaDetalle")
missing_categories_builder.add_header("B", "Rubro")
missing_categories_builder.add_header("C", "Cuenta")
missing_categories_builder.add_header("D", "Detalle")


def from_json_mapper(item_str, key):
    return json.loads(item_str)[key]


def account_details_maper(item_str, _):
    item = json.loads(item_str)

    result = ("%s-%s" % (item['account'], item['details']))
    return result


missing_categories_builder.map_column("A", "account", account_details_maper)
missing_categories_builder.map_column("C", "account", from_json_mapper)
missing_categories_builder.map_column("D", "details", from_json_mapper)

missing_categories_builder.build()

missing_categories_excelwriter.save()
